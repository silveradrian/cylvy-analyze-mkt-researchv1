"""
Enhanced Digital Landscape Excel Exporter with Company Enrichment Data
"""
import io
import json
from datetime import datetime
from typing import Dict, List, Any, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from loguru import logger

from app.core.database import DatabasePool


class EnhancedDigitalLandscapeExporter:
    """Export digital landscape data to Excel format with enriched company data"""
    
    def __init__(self, db: DatabasePool):
        self.db = db
        
    async def export_pipeline_data(self, pipeline_id: str) -> io.BytesIO:
        """Generate comprehensive Excel export for a pipeline"""
        async with self.db.acquire() as conn:
            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # 1. Company DSI Ranking with Enriched Data
            await self._add_enhanced_company_dsi_sheet(wb, conn, pipeline_id)
            
            # 2. Page Level DSI Ranking
            await self._add_page_dsi_sheet(wb, conn, pipeline_id)
            
            # 3. Full Page Level Data
            await self._add_full_page_data_sheet(wb, conn, pipeline_id)
            
            # 4. SERPs & Page Analysis
            await self._add_serp_analysis_sheet(wb, conn, pipeline_id)
            
            # 5. Video Results
            await self._add_video_results_sheet(wb, conn, pipeline_id)
            
            # 6. News Results
            await self._add_news_results_sheet(wb, conn, pipeline_id)
            
            # 7. Dimension Analysis
            await self._add_dimension_analysis_sheet(wb, conn, pipeline_id)
            
            # 8. Company Enrichment Details
            await self._add_company_enrichment_sheet(wb, conn, pipeline_id)
            
            # 9. Pipeline Configuration
            await self._add_pipeline_config_sheet(wb, conn, pipeline_id)
            
            # 10. Summary Dashboard
            await self._add_summary_sheet(wb, conn, pipeline_id)
            
            # Save to BytesIO
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output
    
    async def _add_enhanced_company_dsi_sheet(self, wb: Workbook, conn, pipeline_id: str):
        """Add company DSI ranking sheet with enriched company data"""
        ws = wb.create_sheet("Company DSI Rankings")
        
        # Get company DSI data with enrichment
        companies = await conn.fetch("""
            WITH company_mapping AS (
                -- Map domains to companies
                SELECT DISTINCT
                    cd.domain,
                    cp.id as company_id,
                    cp.company_name,
                    cp.description,
                    cp.industry,
                    cp.employee_count,
                    CASE 
                        WHEN cp.revenue_amount IS NOT NULL THEN 
                            CONCAT(cp.revenue_currency, ' ', cp.revenue_amount::text)
                        ELSE NULL
                    END as annual_revenue,
                    cp.headquarters_location,
                    cp.domain as website,
                    cp.founded_year,
                    cp.source_type as company_type,
                    cp.social_profiles->>'linkedin' as linkedin_url,
                    cp.technologies as tags,
                    NULL as competitors,
                    cd.is_primary
                FROM company_domains cd
                JOIN company_profiles cp ON cd.company_id = cp.id
                WHERE cd.is_active = true
            ),
            company_scores AS (
                SELECT 
                    ds.company_domain,
                    -- Get company name from profiles, fallback to domain extraction
                    COALESCE(
                        cm.company_name,
                        CASE 
                            WHEN ds.company_domain LIKE 'www.%' THEN 
                                INITCAP(SPLIT_PART(SUBSTRING(ds.company_domain FROM 5), '.', 1))
                            ELSE 
                                INITCAP(SPLIT_PART(ds.company_domain, '.', 1))
                        END
                    ) as company_name,
                    ds.dsi_score,
                    ds.keyword_overlap_score,
                    ds.content_relevance_score,
                    ds.market_presence_score,
                    ds.traffic_share_score,
                    ds.metadata,
                    -- Company profile data
                    cm.description,
                    cm.industry,
                    cm.employee_count,
                    cm.annual_revenue,
                    cm.headquarters_location,
                    cm.founded_year,
                    cm.company_type,
                    cm.linkedin_url,
                    cm.tags,
                    cm.competitors,
                    -- Metrics
                    COUNT(DISTINCT sr.keyword_id) as keywords_found,
                    COUNT(DISTINCT sc.url) as pages_scraped,
                    COUNT(DISTINCT oca.id) as pages_analyzed,
                    AVG(sr.position) as avg_position,
                    -- Primary dimensions - get top scoring dimensions
                    array_agg(DISTINCT oda.dimension_name) 
                        FILTER (WHERE oda.dimension_type IN ('persona', 'jtbd_phase') AND CAST(oda.score AS FLOAT) >= 7) as primary_dimensions,
                    -- Additional domains for this company
                    array_agg(DISTINCT other_domains.domain) FILTER (WHERE other_domains.domain != ds.company_domain) as other_domains,
                    -- Extract CTR-based metrics from metadata
                    COALESCE((ds.metadata->>'avg_ctr_percentage')::numeric, 0) as avg_ctr_percentage,
                    COALESCE((ds.metadata->>'weighted_position')::numeric, AVG(sr.position)) as weighted_position,
                    COALESCE((ds.metadata->>'top_3_count')::int, 0) as top_3_positions,
                    COALESCE((ds.metadata->>'top_10_count')::int, 0) as top_10_positions
                FROM dsi_scores ds
                LEFT JOIN company_mapping cm ON cm.domain = ds.company_domain
                LEFT JOIN company_domains other_domains ON other_domains.company_id = cm.company_id
                LEFT JOIN serp_results sr ON sr.domain = ds.company_domain 
                    AND sr.pipeline_execution_id = ds.pipeline_execution_id
                LEFT JOIN scraped_content sc ON sc.domain = ds.company_domain
                    AND sc.pipeline_execution_id = ds.pipeline_execution_id
                LEFT JOIN optimized_content_analysis oca ON oca.url = sc.url
                LEFT JOIN optimized_dimension_analysis oda ON oda.analysis_id = oca.id
                WHERE ds.pipeline_execution_id = $1
                GROUP BY 
                    ds.company_domain, ds.dsi_score, 
                    ds.keyword_overlap_score, ds.content_relevance_score,
                    ds.market_presence_score, ds.traffic_share_score, ds.metadata,
                    cm.company_name, cm.description, cm.industry, cm.employee_count,
                    cm.annual_revenue, cm.headquarters_location, cm.founded_year,
                    cm.company_type, cm.linkedin_url, cm.tags, cm.competitors,
                    ds.metadata->>'avg_ctr_percentage', ds.metadata->>'weighted_position',
                    ds.metadata->>'top_3_count', ds.metadata->>'top_10_count'
            )
            SELECT *,
                   RANK() OVER (ORDER BY dsi_score DESC) as dsi_rank
            FROM company_scores
            ORDER BY dsi_score DESC
        """, pipeline_id)
        
        # Headers
        headers = [
            "Rank", "Company Name", "Primary Domain", "Other Domains", "DSI Score", 
            "Keyword Coverage", "Content Relevance", "Market Presence", "Traffic Share",
            "Avg CTR %", "CTR-Weighted Position", "Top 3 Positions", "Top 10 Positions",
            "Industry", "Description", "Employee Count", "Annual Revenue", 
            "HQ Location", "Founded", "Company Type", "LinkedIn",
            "Keywords Found", "Pages Scraped", "Pages Analyzed", 
            "Avg SERP Position", "Top Scoring Dimensions", "Tags", "Competitors"
        ]
        
        # Style headers
        self._style_headers(ws, headers)
        
        # Add data
        for row_idx, company in enumerate(companies, start=2):
            ws.cell(row=row_idx, column=1, value=company['dsi_rank'])
            ws.cell(row=row_idx, column=2, value=company['company_name'])
            ws.cell(row=row_idx, column=3, value=company['company_domain'])
            ws.cell(row=row_idx, column=4, value=', '.join(company['other_domains']) if company['other_domains'] else '')
            ws.cell(row=row_idx, column=5, value=round(company['dsi_score'], 4))
            ws.cell(row=row_idx, column=6, value=round(company['keyword_overlap_score'], 4))
            ws.cell(row=row_idx, column=7, value=round(company['content_relevance_score'], 4))
            ws.cell(row=row_idx, column=8, value=round(company['market_presence_score'], 4))
            ws.cell(row=row_idx, column=9, value=round(company['traffic_share_score'] or 0, 4))
            ws.cell(row=row_idx, column=10, value=round(float(company['avg_ctr_percentage'] or 0), 2))
            ws.cell(row=row_idx, column=11, value=round(float(company['weighted_position'] or company['avg_position'] or 0), 1))
            ws.cell(row=row_idx, column=12, value=int(company['top_3_positions'] or 0))
            ws.cell(row=row_idx, column=13, value=int(company['top_10_positions'] or 0))
            ws.cell(row=row_idx, column=14, value=company['industry'] or '')
            ws.cell(row=row_idx, column=15, value=(company['description'] or '')[:500])
            ws.cell(row=row_idx, column=16, value=company['employee_count'] or '')
            ws.cell(row=row_idx, column=17, value=company['annual_revenue'] or '')
            ws.cell(row=row_idx, column=18, value=company['headquarters_location'] or '')
            ws.cell(row=row_idx, column=19, value=company['founded_year'] or '')
            ws.cell(row=row_idx, column=20, value=company['company_type'] or '')
            ws.cell(row=row_idx, column=21, value=company['linkedin_url'] or '')
            ws.cell(row=row_idx, column=22, value=company['keywords_found'])
            ws.cell(row=row_idx, column=23, value=company['pages_scraped'])
            ws.cell(row=row_idx, column=24, value=company['pages_analyzed'])
            ws.cell(row=row_idx, column=25, value=round(float(company['avg_position']) if company['avg_position'] else 0, 1))
            ws.cell(row=row_idx, column=26, value=', '.join(company['primary_dimensions'][:3]) if company['primary_dimensions'] else '')
            ws.cell(row=row_idx, column=27, value=self._format_json_field(company['tags']))
            ws.cell(row=row_idx, column=28, value=self._format_json_field(company['competitors']))
            
            # Color code DSI scores
            self._color_score_cell(ws.cell(row=row_idx, column=5), company['dsi_score'])
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    async def _add_page_dsi_sheet(self, wb: Workbook, conn, pipeline_id: str):
        """Add page-level DSI ranking sheet"""
        ws = wb.create_sheet("Page DSI Rankings")
        
        # Get page-level data with CONSISTENT DSI calculations (same formula as company-level)
        pages = await conn.fetch("""
            WITH page_serp_data AS (
                SELECT 
                    sc.url,
                    sc.domain,
                    sc.title,
                    sr.position,
                    sr.serp_type,
                    sr.keyword_id,
                    k.keyword,
                    k.avg_monthly_searches,
                    -- Traffic estimation (same as company-level)
                    COALESCE(k.avg_monthly_searches, 1000) * CASE 
                        WHEN sr.position = 1 THEN 0.2823
                        WHEN sr.position = 2 THEN 0.1572
                        WHEN sr.position = 3 THEN 0.1073
                        WHEN sr.position = 4 THEN 0.0775
                        WHEN sr.position = 5 THEN 0.0588
                        WHEN sr.position = 6 THEN 0.0459
                        WHEN sr.position = 7 THEN 0.0369
                        WHEN sr.position = 8 THEN 0.0302
                        WHEN sr.position = 9 THEN 0.0252
                        WHEN sr.position = 10 THEN 0.0214
                        WHEN sr.position <= 20 THEN 0.0150
                        ELSE 0.0050
                    END as estimated_traffic
                FROM scraped_content sc
                JOIN serp_results sr ON sr.url = sc.url
                JOIN keywords k ON sr.keyword_id = k.id
                WHERE sc.pipeline_execution_id = $1
                    AND sc.status = 'completed'
                    AND sr.result_type = 'organic'
            ),
            page_metrics AS (
                SELECT 
                    url,
                    MAX(domain) as domain,
                    MAX(title) as title,
                    COUNT(DISTINCT keyword_id) as keyword_count,
                    AVG(position) as avg_position,
                    MIN(position) as best_position,
                    SUM(estimated_traffic) as total_estimated_traffic,
                    COUNT(CASE WHEN position <= 10 THEN 1 END) as top_10_count
                FROM page_serp_data
                GROUP BY url
            ),
            content_analysis AS (
                -- Get persona scores for personal relevance (same as company-level)
                SELECT 
                    oca.url,
                    COALESCE(
                        (SELECT AVG(CAST(oda.score AS FLOAT))
                         FROM optimized_dimension_analysis oda 
                         WHERE oda.analysis_id = oca.id 
                         AND oda.dimension_type = 'persona'
                        ), 5.0  -- Default persona score (1-10 scale)
                    ) as persona_score,
                    oca.overall_sentiment,
                    CASE 
                        WHEN oca.key_topics IS NOT NULL AND jsonb_typeof(oca.key_topics) = 'array'
                        THEN array_to_string(ARRAY(SELECT jsonb_array_elements_text(oca.key_topics) LIMIT 5), ', ')
                        ELSE ''
                    END as key_topics_str
                FROM optimized_content_analysis oca
                WHERE oca.project_id IS NULL  -- Pipeline analyses
            ),
            market_totals AS (
                -- Market totals for percentage calculations
                SELECT 
                    COUNT(DISTINCT keyword_id) as total_keywords,
                    SUM(estimated_traffic) as total_market_traffic
                FROM page_serp_data
            )
            SELECT 
                pm.url,
                pm.domain,
                pm.title,
                pm.keyword_count,
                pm.avg_position,
                pm.best_position,
                pm.top_10_count,
                pm.total_estimated_traffic,
                ca.persona_score,
                ca.overall_sentiment,
                ca.key_topics_str,
                -- CONSISTENT DSI COMPONENTS (same as company-level)
                ROUND(
                    (pm.keyword_count::float / mt.total_keywords * 100)::numeric, 
                    2
                ) as keyword_coverage_pct,
                ROUND(
                    (pm.total_estimated_traffic / NULLIF(mt.total_market_traffic, 0) * 100)::numeric,
                    2
                ) as traffic_share_pct,
                -- SAME DSI FORMULA: Keyword Coverage × Share of Traffic × Personal Relevance
                ROUND(
                    (
                        (pm.keyword_count::float / mt.total_keywords * 100) *
                        (pm.total_estimated_traffic / NULLIF(mt.total_market_traffic, 0) * 100) *
                        (ca.persona_score / 10.0)
                    )::numeric,
                    2
                ) as page_dsi_score,
                RANK() OVER (ORDER BY (
                    (pm.keyword_count::float / mt.total_keywords * 100) *
                    (pm.total_estimated_traffic / NULLIF(mt.total_market_traffic, 0) * 100) *
                    (ca.persona_score / 10.0)
                ) DESC) as page_rank
            FROM page_metrics pm
            LEFT JOIN content_analysis ca ON ca.url = pm.url
            CROSS JOIN market_totals mt
            WHERE pm.keyword_count > 0
            ORDER BY page_dsi_score DESC
        """, pipeline_id)
        
        # Headers - Enhanced with comprehensive analysis data
        headers = [
            "Rank", "URL", "Title", "Domain", "Keyword Count", "Avg Position", "Best Position",
            "Page DSI Score", "Keyword Coverage %", "Traffic Share %", "Personal Relevance",
            "Persona Score", "Strategic Imperatives Score", "JTBD Score", "Sentiment", 
            "Key Topics", "Mention Count", "Brand Mentions", "Competitor Mentions", "Overall Insights"
        ]
        
        self._style_headers(ws, headers)
        
        # Add data - Enhanced with comprehensive analysis results
        for row_idx, page in enumerate(pages, start=2):
            ws.cell(row=row_idx, column=1, value=page['page_rank'])
            ws.cell(row=row_idx, column=2, value=page['url'][:255])
            ws.cell(row=row_idx, column=3, value=(page['title'] or '')[:255])
            ws.cell(row=row_idx, column=4, value=page['domain'])
            ws.cell(row=row_idx, column=5, value=page['keyword_count'])
            ws.cell(row=row_idx, column=6, value=round(float(page['avg_position'] or 0), 2))
            ws.cell(row=row_idx, column=7, value=page['best_position'])
            ws.cell(row=row_idx, column=8, value=round(float(page['page_dsi_score'] or 0), 4))
            ws.cell(row=row_idx, column=9, value=round(float(page['keyword_coverage_pct'] or 0), 2))
            ws.cell(row=row_idx, column=10, value=round(float(page['traffic_share_pct'] or 0), 4))
            ws.cell(row=row_idx, column=11, value=round(float(page['persona_score'] or 0), 2))
            ws.cell(row=row_idx, column=12, value=round(float(page['persona_score'] or 0), 2))
            ws.cell(row=row_idx, column=13, value=round(float(page['strategic_imperative_score'] or 0), 2))
            ws.cell(row=row_idx, column=14, value=round(float(page['jtbd_score'] or 0), 2))
            ws.cell(row=row_idx, column=15, value=page['overall_sentiment'] or '')
            ws.cell(row=row_idx, column=16, value=page['key_topics_str'] or '')
            ws.cell(row=row_idx, column=17, value=page.get('mention_count', 0))
            ws.cell(row=row_idx, column=18, value=page.get('brand_mention_count', 0))
            ws.cell(row=row_idx, column=19, value=page.get('competitor_mention_count', 0))
            ws.cell(row=row_idx, column=20, value=(page.get('overall_insights', '') or '')[:255])
            
            self._color_score_cell(ws.cell(row=row_idx, column=8), float(page['page_dsi_score'] or 0))
        
        self._auto_adjust_columns(ws)
    
    async def _add_full_page_data_sheet(self, wb: Workbook, conn, pipeline_id: str):
        """Add comprehensive page-level data"""
        ws = wb.create_sheet("Full Page Data")
        
        # Get all page data with analysis
        pages = await conn.fetch("""
            SELECT 
                sc.url,
                sc.domain,
                sc.title,
                sc.meta_description,
                LENGTH(sc.content) as content_length,
                sc.status as scrape_status,
                sc.created_at as scraped_at,
                sr.keyword_id,
                k.keyword,
                sr.position,
                sr.serp_type,
                sr.snippet,
                oca.overall_insights,
                oca.overall_sentiment,
                CASE 
                    WHEN oca.key_topics IS NOT NULL AND jsonb_typeof(oca.key_topics) = 'array'
                    THEN array_to_string(ARRAY(SELECT jsonb_array_elements_text(oca.key_topics)), ', ')
                    ELSE ''
                END as key_topics_str,
                oca.mentions,
                -- Get all dimension scores as JSON
                (
                    SELECT json_object_agg(
                        oda.dimension_name || '_' || oda.dimension_type, 
                        json_build_object(
                            'score', oda.score, 
                            'evidence', oda.key_evidence,
                            'confidence', oda.confidence,
                            'is_primary', false
                        )
                    )
                    FROM optimized_dimension_analysis oda
                    WHERE oda.analysis_id = oca.id
                ) as dimension_scores
            FROM scraped_content sc
            JOIN serp_results sr ON sr.url = sc.url AND sr.pipeline_execution_id = sc.pipeline_execution_id
            JOIN keywords k ON sr.keyword_id = k.id
            LEFT JOIN optimized_content_analysis oca ON oca.url = sc.url
            WHERE sc.pipeline_execution_id = $1
            ORDER BY k.keyword, sr.position
        """, pipeline_id)
        
        # Headers
        headers = [
            "URL", "Domain", "Title", "Meta Description", "Content Length",
            "Scrape Status", "Scraped At", "Keyword", "Position", "SERP Type",
            "Snippet", "Overall Insights", "Sentiment", "Key Topics", "Mentions",
            "Lending Leader Score", "Banking Architect Score", "Payments Innovator Score",
            "Problem ID Score", "Solution Explore Score", "Requirements Score",
            "Supplier Select Score", "Validation Score", "Primary Dimensions"
        ]
        
        self._style_headers(ws, headers)
        
        # Add data
        for row_idx, page in enumerate(pages, start=2):
            dim_scores = page['dimension_scores'] or {}
            
            # Handle case where dim_scores might be JSON string
            if isinstance(dim_scores, str):
                try:
                    import json
                    dim_scores = json.loads(dim_scores)
                except:
                    dim_scores = {}
            
            ws.cell(row=row_idx, column=1, value=page['url'][:255])
            ws.cell(row=row_idx, column=2, value=page['domain'])
            ws.cell(row=row_idx, column=3, value=(page['title'] or '')[:255])
            ws.cell(row=row_idx, column=4, value=(page['meta_description'] or '')[:255])
            ws.cell(row=row_idx, column=5, value=page['content_length'] or 0)
            ws.cell(row=row_idx, column=6, value=page['scrape_status'])
            ws.cell(row=row_idx, column=7, value=page['scraped_at'].strftime('%Y-%m-%d %H:%M') if page['scraped_at'] else '')
            ws.cell(row=row_idx, column=8, value=page['keyword'])
            ws.cell(row=row_idx, column=9, value=page['position'])
            ws.cell(row=row_idx, column=10, value=page['serp_type'])
            ws.cell(row=row_idx, column=11, value=(page['snippet'] or '')[:255])
            ws.cell(row=row_idx, column=12, value=(page['overall_insights'] or '')[:500])
            ws.cell(row=row_idx, column=13, value=page['overall_sentiment'] or '')
            ws.cell(row=row_idx, column=14, value=page['key_topics_str'] or '')
            ws.cell(row=row_idx, column=15, value=json.dumps(page['mentions'])[:255] if page['mentions'] else '')
            
            # Persona scores
            ws.cell(row=row_idx, column=16, value=self._get_dimension_score(dim_scores, 'The Modern Lending Leader_persona'))
            ws.cell(row=row_idx, column=17, value=self._get_dimension_score(dim_scores, 'The Digital Banking Architect_persona'))
            ws.cell(row=row_idx, column=18, value=self._get_dimension_score(dim_scores, 'The Payments Innovator_persona'))
            
            # JTBD scores
            ws.cell(row=row_idx, column=19, value=self._get_dimension_score(dim_scores, 'Problem Identification_jtbd_phase'))
            ws.cell(row=row_idx, column=20, value=self._get_dimension_score(dim_scores, 'Solution Exploration_jtbd_phase'))
            ws.cell(row=row_idx, column=21, value=self._get_dimension_score(dim_scores, 'Requirements Building_jtbd_phase'))
            ws.cell(row=row_idx, column=22, value=self._get_dimension_score(dim_scores, 'Supplier Selection_jtbd_phase'))
            ws.cell(row=row_idx, column=23, value=self._get_dimension_score(dim_scores, 'Validation_jtbd_phase'))
            
            # Primary dimensions
            primary_dims = [k.split('_')[0] for k, v in dim_scores.items() if isinstance(v, dict) and v.get('is_primary')]
            ws.cell(row=row_idx, column=24, value=', '.join(primary_dims))
        
        self._auto_adjust_columns(ws)
    
    async def _add_dimension_analysis_sheet(self, wb: Workbook, conn, pipeline_id: str):
        """Add dimension analysis with primary dimension indicators"""
        ws = wb.create_sheet("Dimension Analysis")
        
        # Get dimension performance data
        dimensions = await conn.fetch("""
            WITH dimension_stats AS (
                SELECT 
                    oda.dimension_id,
                    oda.dimension_name,
                    oda.dimension_type,
                    COUNT(*) as usage_count,
                    AVG(CAST(oda.score AS FLOAT)) as avg_score,
                    STDDEV(CAST(oda.score AS FLOAT)) as score_stddev,
                    MIN(CAST(oda.score AS FLOAT)) as min_score,
                    MAX(CAST(oda.score AS FLOAT)) as max_score,
                    COUNT(CASE WHEN CAST(oda.score AS INT) >= 7 THEN 1 END) as high_scores,
                    COUNT(CASE WHEN CAST(oda.score AS INT) <= 3 THEN 1 END) as low_scores,
                    -- Group dimensions by type
                    ARRAY[oda.dimension_type] as dimension_groups,
                    -- Mark as primary if it's a high-scoring dimension
                    bool_or(CAST(oda.score AS FLOAT) >= 8) as is_primary_dimension
                FROM optimized_dimension_analysis oda
                JOIN optimized_content_analysis oca ON oca.id = oda.analysis_id
                JOIN scraped_content sc ON sc.url = oca.url
                WHERE sc.pipeline_execution_id = $1
                GROUP BY oda.dimension_id, oda.dimension_name, oda.dimension_type
            )
            SELECT *
            FROM dimension_stats
            ORDER BY dimension_type, avg_score DESC
        """, pipeline_id)
        
        # Headers
        headers = [
            "Dimension Name", "Type", "Is Primary", "Dimension Groups",
            "Usage Count", "Avg Score", "Std Dev", "Min Score", "Max Score",
            "High Scores (7+)", "Low Scores (≤3)", "Performance"
        ]
        
        self._style_headers(ws, headers)
        
        # Add data
        for row_idx, dim in enumerate(dimensions, start=2):
            ws.cell(row=row_idx, column=1, value=dim['dimension_name'])
            ws.cell(row=row_idx, column=2, value=dim['dimension_type'])
            ws.cell(row=row_idx, column=3, value='Yes' if dim['is_primary_dimension'] else 'No')
            ws.cell(row=row_idx, column=4, value=', '.join(dim['dimension_groups']) if dim['dimension_groups'] else '')
            ws.cell(row=row_idx, column=5, value=dim['usage_count'])
            ws.cell(row=row_idx, column=6, value=round(float(dim['avg_score']), 2))
            ws.cell(row=row_idx, column=7, value=round(float(dim['score_stddev'] or 0), 2))
            ws.cell(row=row_idx, column=8, value=round(float(dim['min_score']), 1))
            ws.cell(row=row_idx, column=9, value=round(float(dim['max_score']), 1))
            ws.cell(row=row_idx, column=10, value=dim['high_scores'])
            ws.cell(row=row_idx, column=11, value=dim['low_scores'])
            
            # Performance indicator
            avg_score = float(dim['avg_score'])
            if avg_score >= 6:
                performance = 'Overperforming'
                color = '90EE90'
            elif avg_score >= 4:
                performance = 'Normal'
                color = 'FFFFE0'
            else:
                performance = 'Underperforming'
                color = 'FFB6C1'
            
            perf_cell = ws.cell(row=row_idx, column=12, value=performance)
            perf_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            # Highlight primary dimensions
            if dim['is_primary_dimension']:
                for col in range(1, 4):
                    ws.cell(row=row_idx, column=col).font = Font(bold=True)
        
        self._auto_adjust_columns(ws)
    
    async def _add_serp_analysis_sheet(self, wb: Workbook, conn, pipeline_id: str):
        """Add SERP analysis sheet"""
        ws = wb.create_sheet("SERP Analysis")
        
        # Get SERP analysis data
        serp_data = await conn.fetch("""
            WITH serp_analysis AS (
                SELECT 
                    k.keyword,
                    sr.serp_type,
                    sr.position,
                    sr.url,
                    sr.domain,
                    sr.title,
                    sr.snippet,
                    sc.status as scrape_status,
                    oca.id IS NOT NULL as is_analyzed,
                    ds.dsi_score,
                    COALESCE(
                        cp.company_name,
                        CASE 
                            WHEN sr.domain LIKE 'www.%' THEN 
                                INITCAP(SPLIT_PART(SUBSTRING(sr.domain FROM 5), '.', 1))
                            ELSE 
                                INITCAP(SPLIT_PART(sr.domain, '.', 1))
                        END
                    ) as company_name
                FROM serp_results sr
                JOIN keywords k ON sr.keyword_id = k.id
                LEFT JOIN scraped_content sc ON sc.url = sr.url 
                    AND sc.pipeline_execution_id = sr.pipeline_execution_id
                LEFT JOIN optimized_content_analysis oca ON oca.url = sr.url
                LEFT JOIN dsi_scores ds ON ds.company_domain = sr.domain
                    AND ds.pipeline_execution_id = sr.pipeline_execution_id
                LEFT JOIN company_domains cd ON cd.domain = sr.domain AND cd.is_active = true
                LEFT JOIN company_profiles cp ON cp.id = cd.company_id
                WHERE sr.pipeline_execution_id = $1
            )
            SELECT *,
                   RANK() OVER (PARTITION BY keyword ORDER BY position) as keyword_rank
            FROM serp_analysis
            ORDER BY keyword, position
        """, pipeline_id)
        
        # Headers
        headers = [
            "Keyword", "Rank", "SERP Type", "Position", "URL", "Company Name", "Domain",
            "Title", "Snippet", "Scrape Status", "Analyzed", "Company DSI Score"
        ]
        
        self._style_headers(ws, headers)
        
        # Add data
        for row_idx, serp in enumerate(serp_data, start=2):
            ws.cell(row=row_idx, column=1, value=serp['keyword'])
            ws.cell(row=row_idx, column=2, value=serp['keyword_rank'])
            ws.cell(row=row_idx, column=3, value=serp['serp_type'])
            ws.cell(row=row_idx, column=4, value=serp['position'])
            ws.cell(row=row_idx, column=5, value=serp['url'][:255])
            ws.cell(row=row_idx, column=6, value=serp['company_name'])
            ws.cell(row=row_idx, column=7, value=serp['domain'])
            ws.cell(row=row_idx, column=8, value=(serp['title'] or '')[:255])
            ws.cell(row=row_idx, column=9, value=(serp['snippet'] or '')[:500])
            ws.cell(row=row_idx, column=10, value=serp['scrape_status'] or 'Not scraped')
            ws.cell(row=row_idx, column=11, value='Yes' if serp['is_analyzed'] else 'No')
            ws.cell(row=row_idx, column=12, value=round(float(serp['dsi_score']), 4) if serp['dsi_score'] else '')
            
            # Highlight top positions
            if serp['position'] <= 3:
                ws.cell(row=row_idx, column=4).fill = PatternFill(
                    start_color="90EE90", end_color="90EE90", fill_type="solid"
                )
        
        self._auto_adjust_columns(ws)
    
    async def _add_video_results_sheet(self, wb: Workbook, conn, pipeline_id: str):
        """Add video results sheet"""
        ws = wb.create_sheet("Video Results")
        
        # Get video results
        videos = await conn.fetch("""
            SELECT 
                sr.url,
                sr.domain,
                sr.title,
                sr.snippet,
                sr.position,
                k.keyword,
                sc.status as scrape_status,
                ds.dsi_score,
                ycc.channel_id,
                vs.channel_title as channel_name,
                ycc.company_domain as resolved_company_domain,
                ycc.company_name as resolved_company_name,
                ycc.confidence_score,
                vs.view_count,
                vs.like_count,
                vs.comment_count,
                vs.duration_seconds,
                vs.published_at,
                COALESCE(
                    cp.company_name,
                    ycc.company_name,
                    CASE 
                        WHEN COALESCE(ycc.company_domain, sr.domain) LIKE 'www.%' THEN 
                            INITCAP(SPLIT_PART(SUBSTRING(COALESCE(ycc.company_domain, sr.domain) FROM 5), '.', 1))
                        ELSE 
                            INITCAP(SPLIT_PART(COALESCE(ycc.company_domain, sr.domain), '.', 1))
                    END
                ) as company_name
            FROM serp_results sr
            JOIN keywords k ON sr.keyword_id = k.id
            LEFT JOIN scraped_content sc ON sc.url = sr.url 
                AND sc.pipeline_execution_id = sr.pipeline_execution_id
            LEFT JOIN dsi_scores ds ON ds.company_domain = sr.domain
                AND ds.pipeline_execution_id = sr.pipeline_execution_id
            LEFT JOIN video_snapshots vs ON vs.video_url = sr.url
            LEFT JOIN youtube_channel_companies ycc ON ycc.channel_id = vs.channel_id
            LEFT JOIN company_domains cd ON cd.domain = COALESCE(ycc.company_domain, sr.domain) AND cd.is_active = true
            LEFT JOIN company_profiles cp ON cp.id = cd.company_id
            WHERE sr.pipeline_execution_id = $1
                AND sr.serp_type = 'video'
            ORDER BY k.keyword, sr.position
        """, pipeline_id)
        
        # Headers
        headers = [
            "Keyword", "Position", "URL", "Company Name", "Domain", "Title", "Description",
            "DSI Score", "Channel ID", "Channel Name", "Resolved Company Domain", 
            "Resolved Company Name", "Confidence Score", 
            "View Count", "Like Count", "Comment Count", "Duration (seconds)", "Published Date",
            "Scrape Status"
        ]
        
        self._style_headers(ws, headers)
        
        # Add data
        for row_idx, video in enumerate(videos, start=2):
            ws.cell(row=row_idx, column=1, value=video['keyword'])
            ws.cell(row=row_idx, column=2, value=video['position'])
            ws.cell(row=row_idx, column=3, value=video['url'][:255])
            ws.cell(row=row_idx, column=4, value=video['company_name'])
            ws.cell(row=row_idx, column=5, value=video['domain'])
            ws.cell(row=row_idx, column=6, value=(video['title'] or '')[:255])
            ws.cell(row=row_idx, column=7, value=(video['snippet'] or '')[:500])
            ws.cell(row=row_idx, column=8, value=round(float(video['dsi_score']), 4) if video['dsi_score'] else '')
            ws.cell(row=row_idx, column=9, value=video['channel_id'] or '')
            ws.cell(row=row_idx, column=10, value=video['channel_name'] or '')
            ws.cell(row=row_idx, column=11, value=video['resolved_company_domain'] or '')
            ws.cell(row=row_idx, column=12, value=video['resolved_company_name'] or '')
            ws.cell(row=row_idx, column=13, value=round(float(video['confidence_score']), 2) if video['confidence_score'] else '')
            ws.cell(row=row_idx, column=14, value=video['view_count'] or 0)
            ws.cell(row=row_idx, column=15, value=video['like_count'] or 0)
            ws.cell(row=row_idx, column=16, value=video['comment_count'] or 0)
            ws.cell(row=row_idx, column=17, value=video['duration_seconds'] or 0)
            ws.cell(row=row_idx, column=18, value=video['published_at'].strftime('%Y-%m-%d') if video['published_at'] else '')
            ws.cell(row=row_idx, column=19, value=video['scrape_status'] or 'Not scraped')
        
        self._auto_adjust_columns(ws)
    
    async def _add_news_results_sheet(self, wb: Workbook, conn, pipeline_id: str):
        """Add news results sheet"""
        ws = wb.create_sheet("News Results")
        
        # Get news results
        news = await conn.fetch("""
            SELECT 
                sr.url,
                sr.domain,
                sr.title,
                sr.snippet,
                sr.position,
                k.keyword,
                sr.search_date,
                sc.status as scrape_status,
                sc.content IS NOT NULL as has_content,
                oca.overall_sentiment,
                CASE 
                    WHEN oca.key_topics IS NOT NULL AND jsonb_typeof(oca.key_topics) = 'array'
                    THEN array_to_string(ARRAY(SELECT jsonb_array_elements_text(oca.key_topics) LIMIT 5), ', ')
                    ELSE ''
                END as key_topics_str,
                ds.dsi_score,
                COALESCE(
                    cp.company_name,
                    CASE 
                        WHEN sr.domain LIKE 'www.%' THEN 
                            INITCAP(SPLIT_PART(SUBSTRING(sr.domain FROM 5), '.', 1))
                        ELSE 
                            INITCAP(SPLIT_PART(sr.domain, '.', 1))
                    END
                ) as company_name
            FROM serp_results sr
            JOIN keywords k ON sr.keyword_id = k.id
            LEFT JOIN scraped_content sc ON sc.url = sr.url 
                AND sc.pipeline_execution_id = sr.pipeline_execution_id
            LEFT JOIN optimized_content_analysis oca ON oca.url = sr.url
            LEFT JOIN dsi_scores ds ON ds.company_domain = sr.domain
                AND ds.pipeline_execution_id = sr.pipeline_execution_id
            LEFT JOIN company_domains cd ON cd.domain = sr.domain AND cd.is_active = true
            LEFT JOIN company_profiles cp ON cp.id = cd.company_id
            WHERE sr.pipeline_execution_id = $1
                AND sr.serp_type = 'news'
            ORDER BY sr.search_date DESC, k.keyword, sr.position
        """, pipeline_id)
        
        # Headers
        headers = [
            "Date", "Keyword", "Position", "URL", "Company Name", "Domain", "Title", "Snippet",
            "DSI Score", "Scrape Status", "Has Content", "Sentiment", "Key Topics"
        ]
        
        self._style_headers(ws, headers)
        
        # Add data
        for row_idx, article in enumerate(news, start=2):
            ws.cell(row=row_idx, column=1, value=article['search_date'].strftime('%Y-%m-%d'))
            ws.cell(row=row_idx, column=2, value=article['keyword'])
            ws.cell(row=row_idx, column=3, value=article['position'])
            ws.cell(row=row_idx, column=4, value=article['url'][:255])
            ws.cell(row=row_idx, column=5, value=article['company_name'])
            ws.cell(row=row_idx, column=6, value=article['domain'])
            ws.cell(row=row_idx, column=7, value=(article['title'] or '')[:255])
            ws.cell(row=row_idx, column=8, value=(article['snippet'] or '')[:500])
            ws.cell(row=row_idx, column=9, value=round(float(article['dsi_score']), 4) if article['dsi_score'] else '')
            ws.cell(row=row_idx, column=10, value=article['scrape_status'] or 'Not scraped')
            ws.cell(row=row_idx, column=11, value='Yes' if article['has_content'] else 'No')
            ws.cell(row=row_idx, column=12, value=article['overall_sentiment'] or '')
            ws.cell(row=row_idx, column=13, value=article['key_topics_str'] or '')
        
        self._auto_adjust_columns(ws)
    
    async def _add_summary_sheet(self, wb: Workbook, conn, pipeline_id: str):
        """Add summary dashboard as first sheet"""
        ws = wb.create_sheet("Summary Dashboard", 0)
        
        # Get summary statistics
        stats = await conn.fetchrow("""
            WITH pipeline_stats AS (
                SELECT 
                    pe.created_at,
                    pe.completed_at,
                    pe.status,
                    (SELECT COUNT(DISTINCT keyword_id) FROM serp_results WHERE pipeline_execution_id = $1) as total_keywords,
                    (SELECT COUNT(*) FROM serp_results WHERE pipeline_execution_id = $1) as total_serp_results,
                    (SELECT COUNT(*) FROM serp_results WHERE pipeline_execution_id = $1 AND serp_type = 'organic') as organic_results,
                    (SELECT COUNT(*) FROM serp_results WHERE pipeline_execution_id = $1 AND serp_type = 'news') as news_results,
                    (SELECT COUNT(*) FROM serp_results WHERE pipeline_execution_id = $1 AND serp_type = 'video') as video_results,
                    (SELECT COUNT(DISTINCT domain) FROM serp_results WHERE pipeline_execution_id = $1) as unique_domains,
                    (SELECT COUNT(*) FROM scraped_content WHERE pipeline_execution_id = $1) as total_scraped,
                    (SELECT COUNT(*) FROM scraped_content WHERE pipeline_execution_id = $1 AND status = 'completed') as successful_scrapes,
                    (SELECT COUNT(DISTINCT oca.url) FROM optimized_content_analysis oca 
                     JOIN scraped_content sc ON oca.url = sc.url 
                     WHERE sc.pipeline_execution_id = $1) as total_analyzed,
                    (SELECT COUNT(*) FROM dsi_scores WHERE pipeline_execution_id = $1) as companies_scored,
                    (SELECT AVG(dsi_score) FROM dsi_scores WHERE pipeline_execution_id = $1) as avg_dsi_score,
                    (SELECT MAX(dsi_score) FROM dsi_scores WHERE pipeline_execution_id = $1) as max_dsi_score,
                    (SELECT COUNT(DISTINCT domain) FROM serp_results WHERE pipeline_execution_id = $1 AND domain IS NOT NULL) as companies_enriched
                FROM pipeline_executions pe
                WHERE pe.id = $1
            )
            SELECT * FROM pipeline_stats
        """, pipeline_id)
        
        # Title
        ws['A1'] = "Digital Landscape Analysis Report"
        ws['A1'].font = Font(size=18, bold=True)
        ws['A2'] = f"Pipeline ID: {pipeline_id}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Pipeline Info
        ws['A5'] = "Pipeline Information"
        ws['A5'].font = Font(size=14, bold=True)
        
        info_data = [
            ['Status:', stats['status']],
            ['Started:', stats['created_at'].strftime('%Y-%m-%d %H:%M') if stats['created_at'] else 'N/A'],
            ['Completed:', stats['completed_at'].strftime('%Y-%m-%d %H:%M') if stats['completed_at'] else 'In Progress'],
            ['Duration:', self._calculate_duration(stats['created_at'], stats['completed_at'])],
        ]
        
        for idx, (label, value) in enumerate(info_data, start=6):
            ws[f'A{idx}'] = label
            ws[f'B{idx}'] = value
            ws[f'A{idx}'].font = Font(bold=True)
        
        # Key Metrics
        ws['D5'] = "Key Metrics"
        ws['D5'].font = Font(size=14, bold=True)
        
        metrics_data = [
            ['Total Keywords:', stats['total_keywords']],
            ['SERP Results:', stats['total_serp_results']],
            ['Unique Domains:', stats['unique_domains']],
            ['Companies Enriched:', stats['companies_enriched'] or 0],
            ['Pages Scraped:', f"{stats['successful_scrapes']}/{stats['total_scraped']} ({stats['successful_scrapes']/stats['total_scraped']*100:.1f}%)"],
            ['Pages Analyzed:', stats['total_analyzed']],
            ['Companies Scored:', stats['companies_scored']],
            ['Average DSI Score:', f"{float(stats['avg_dsi_score'] or 0):.3f}"],
            ['Top DSI Score:', f"{float(stats['max_dsi_score'] or 0):.3f}"],
        ]
        
        for idx, (label, value) in enumerate(metrics_data, start=6):
            ws[f'D{idx}'] = label
            ws[f'E{idx}'] = value
            ws[f'D{idx}'].font = Font(bold=True)
        
        # SERP Breakdown
        ws['A16'] = "SERP Results Breakdown"
        ws['A16'].font = Font(size=14, bold=True)
        
        serp_headers = ['Type', 'Count', 'Percentage']
        for col_idx, header in enumerate(serp_headers, start=1):
            cell = ws.cell(row=17, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        serp_data = [
            ['Organic', stats['organic_results'], f"{stats['organic_results']/stats['total_serp_results']*100:.1f}%"],
            ['News', stats['news_results'], f"{stats['news_results']/stats['total_serp_results']*100:.1f}%"],
            ['Video', stats['video_results'], f"{stats['video_results']/stats['total_serp_results']*100:.1f}%"],
        ]
        
        for row_idx, row_data in enumerate(serp_data, start=18):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Top Companies by DSI
        ws['D16'] = "Top 10 Companies by DSI Score"
        ws['D16'].font = Font(size=14, bold=True)
        
        top_companies = await conn.fetch("""
            SELECT 
                COALESCE(
                    cp.company_name,
                    CASE 
                        WHEN ds.company_domain LIKE 'www.%' THEN 
                            INITCAP(SPLIT_PART(SUBSTRING(ds.company_domain FROM 5), '.', 1))
                        ELSE 
                            INITCAP(SPLIT_PART(ds.company_domain, '.', 1))
                    END
                ) as company_name,
                ds.company_domain, 
                ds.dsi_score
            FROM dsi_scores ds
            LEFT JOIN company_domains cd ON cd.domain = ds.company_domain AND cd.is_active = true
            LEFT JOIN company_profiles cp ON cp.id = cd.company_id
            WHERE ds.pipeline_execution_id = $1
            ORDER BY ds.dsi_score DESC
            LIMIT 10
        """, pipeline_id)
        
        company_headers = ['Rank', 'Company', 'Domain', 'DSI Score']
        for col_idx, header in enumerate(company_headers, start=4):
            cell = ws.cell(row=17, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for row_idx, company in enumerate(top_companies, start=18):
            ws.cell(row=row_idx, column=4, value=row_idx - 17)
            ws.cell(row=row_idx, column=5, value=company['company_name'])
            ws.cell(row=row_idx, column=6, value=company['company_domain'])
            ws.cell(row=row_idx, column=7, value=round(float(company['dsi_score']), 4))
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def _style_headers(self, ws, headers: List[str]):
        """Apply consistent header styling"""
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    def _color_score_cell(self, cell, score: float):
        """Color code score cells"""
        if score >= 0.7:
            color = "90EE90"  # Light green
        elif score >= 0.4:
            color = "FFFFE0"  # Light yellow
        else:
            color = "FFB6C1"  # Light red
        
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = None
            
            for cell in column:
                try:
                    # Skip merged cells
                    if hasattr(cell, 'column_letter'):
                        if not column_letter:
                            column_letter = cell.column_letter
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            if column_letter:
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _get_dimension_score(self, dim_scores: Dict, dimension_key: str) -> float:
        """Extract dimension score from JSON"""
        if not dim_scores:
            return 0
        
        # Handle case where dim_scores might be JSON string
        if isinstance(dim_scores, str):
            try:
                import json
                dim_scores = json.loads(dim_scores)
            except:
                return 0
        
        if dimension_key in dim_scores:
            score_data = dim_scores[dimension_key]
            if isinstance(score_data, dict):
                return float(score_data.get('score', 0))
            else:
                return float(score_data or 0)
        return 0
    
    def _format_json_field(self, field) -> str:
        """Format JSON fields for display"""
        if not field:
            return ''
        
        # Handle JSONB fields
        if isinstance(field, str):
            try:
                field = json.loads(field)
            except:
                return field[:255]
        
        if isinstance(field, list):
            return ', '.join(str(item) for item in field[:5])
        elif isinstance(field, dict):
            return json.dumps(field)[:255]
        else:
            return str(field)[:255]
    
    def _parse_metadata(self, metadata) -> Dict:
        """Parse metadata field from JSON"""
        if not metadata:
            return {}
        
        if isinstance(metadata, str):
            try:
                return json.loads(metadata)
            except:
                return {}
        
        return metadata if isinstance(metadata, dict) else {}
    
    def _calculate_duration(self, start_time, end_time) -> str:
        """Calculate duration between timestamps"""
        if not start_time:
            return "N/A"
        
        end = end_time or datetime.now(start_time.tzinfo)
        duration = end - start_time
        
        hours = duration.total_seconds() / 3600
        if hours < 1:
            return f"{int(duration.total_seconds() / 60)} minutes"
        elif hours < 24:
            return f"{hours:.1f} hours"
        else:
            return f"{hours / 24:.1f} days"
    
    async def _add_company_enrichment_sheet(self, wb: Workbook, conn, pipeline_id: str):
        """Add detailed company enrichment data"""
        ws = wb.create_sheet("Company Enrichment Details")
        
        # Get all enriched companies with full details
        companies = await conn.fetch("""
            SELECT DISTINCT
                cp.*,
                cd.domain,
                cd.domain_type,
                cd.is_primary,
                -- Extract individual social profiles
                cp.social_profiles->>'linkedin' as linkedin_url,
                cp.social_profiles->>'twitter' as twitter_url,
                cp.social_profiles->>'facebook' as facebook_url,
                -- Count of associated pages
                (
                    SELECT COUNT(DISTINCT sr.url)
                    FROM serp_results sr
                    WHERE sr.domain = cd.domain
                    AND sr.pipeline_execution_id = $1
                ) as page_count,
                -- YouTube presence
                (
                    SELECT COUNT(DISTINCT ycc.channel_id)
                    FROM youtube_channel_companies ycc
                    WHERE ycc.company_domain = cd.domain
                ) as youtube_channels
            FROM company_profiles cp
            JOIN company_domains cd ON cd.company_id = cp.id
            WHERE cd.domain IN (
                SELECT DISTINCT domain 
                FROM serp_results 
                WHERE pipeline_execution_id = $1
            )
            ORDER BY cp.company_name, cd.is_primary DESC
        """, pipeline_id)
        
        # Headers
        headers = [
            "Company Name", "Domain", "Domain Type", "Is Primary", "Industry", "Sub-Industry",
            "Description", "Employee Count", "Revenue", "Founded Year",
            "Headquarters", "Website", "Source", "Source Type",
            "LinkedIn", "Twitter", "Facebook", "Technologies",
            "Page Count", "YouTube Channels", "Created At", "Updated At"
        ]
        
        self._style_headers(ws, headers)
        
        # Add data
        for row_idx, company in enumerate(companies, start=2):
            ws.cell(row=row_idx, column=1, value=company['company_name'])
            ws.cell(row=row_idx, column=2, value=company['domain'])
            ws.cell(row=row_idx, column=3, value=company['domain_type'])
            ws.cell(row=row_idx, column=4, value="Yes" if company['is_primary'] else "No")
            ws.cell(row=row_idx, column=5, value=company['industry'] or '')
            ws.cell(row=row_idx, column=6, value=company['sub_industry'] or '')
            ws.cell(row=row_idx, column=7, value=(company['description'] or '')[:500])
            ws.cell(row=row_idx, column=8, value=company['employee_count'] or '')
            
            # Format revenue
            revenue = ''
            if company['revenue_amount']:
                revenue = f"{company['revenue_currency'] or 'USD'} {company['revenue_amount']:,.0f}"
            ws.cell(row=row_idx, column=9, value=revenue)
            
            ws.cell(row=row_idx, column=10, value=company['founded_year'] or '')
            
            # Format headquarters
            hq = ''
            if company['headquarters_location']:
                hq_data = json.loads(company['headquarters_location']) if isinstance(company['headquarters_location'], str) else company['headquarters_location']
                if hq_data:
                    hq = f"{hq_data.get('city', '')}, {hq_data.get('country', '')}".strip(', ')
            ws.cell(row=row_idx, column=11, value=hq)
            
            ws.cell(row=row_idx, column=12, value=company.get('website', company['domain']) or '')
            ws.cell(row=row_idx, column=13, value=company['source'] or '')
            ws.cell(row=row_idx, column=14, value=company['source_type'] or '')
            ws.cell(row=row_idx, column=15, value=company['linkedin_url'] or '')
            ws.cell(row=row_idx, column=16, value=company['twitter_url'] or '')
            ws.cell(row=row_idx, column=17, value=company['facebook_url'] or '')
            
            # Format technologies
            tech = ''
            if company['technologies']:
                tech_list = json.loads(company['technologies']) if isinstance(company['technologies'], str) else company['technologies']
                tech = ', '.join(tech_list[:10])  # Limit to first 10
            ws.cell(row=row_idx, column=18, value=tech)
            
            ws.cell(row=row_idx, column=19, value=company['page_count'] or 0)
            ws.cell(row=row_idx, column=20, value=company['youtube_channels'] or 0)
            ws.cell(row=row_idx, column=21, value=company['created_at'].strftime('%Y-%m-%d %H:%M') if company['created_at'] else '')
            ws.cell(row=row_idx, column=22, value=company['updated_at'].strftime('%Y-%m-%d %H:%M') if company['updated_at'] else '')
        
        self._auto_adjust_columns(ws)
    
    async def _add_pipeline_config_sheet(self, wb: Workbook, conn, pipeline_id: str):
        """Add pipeline configuration and execution details"""
        ws = wb.create_sheet("Pipeline Configuration")
        
        # Get pipeline details
        pipeline = await conn.fetchrow("""
            SELECT 
                pe.*,
                (SELECT COUNT(DISTINCT keyword_id) FROM serp_results WHERE pipeline_execution_id = pe.id) as keyword_count,
                COUNT(DISTINCT pps.phase_name) as phase_count,
                MIN(pps.started_at) as pipeline_started,
                MAX(pps.completed_at) as pipeline_completed,
                EXTRACT(EPOCH FROM (MAX(pps.completed_at) - MIN(pps.started_at)))/60 as total_runtime_minutes
            FROM pipeline_executions pe
            LEFT JOIN pipeline_phase_status pps ON pps.pipeline_execution_id = pe.id
            WHERE pe.id = $1
            GROUP BY pe.id
        """, pipeline_id)
        
        # Get phase details
        phases = await conn.fetch("""
            SELECT 
                phase_name,
                status,
                started_at,
                completed_at,
                EXTRACT(EPOCH FROM (completed_at - started_at))/60 as runtime_minutes
            FROM pipeline_phase_status
            WHERE pipeline_execution_id = $1
            ORDER BY started_at
        """, pipeline_id)
        
        # Get keywords
        keywords = await conn.fetch("""
            SELECT DISTINCT k.keyword
            FROM keywords k
            JOIN serp_results sr ON sr.keyword_id = k.id
            WHERE sr.pipeline_execution_id = $1
            ORDER BY k.keyword
        """, pipeline_id)
        
        # Get regions
        regions = await conn.fetch("""
            SELECT DISTINCT sr.location
            FROM serp_results sr
            WHERE sr.pipeline_execution_id = $1
            AND sr.location IS NOT NULL
        """, pipeline_id)
        
        # Pipeline overview section
        ws.merge_cells('A1:B1')
        ws['A1'] = 'Pipeline Overview'
        ws['A1'].font = Font(bold=True, size=14)
        
        overview_data = [
            ('Pipeline ID', str(pipeline['id'])),
            ('Status', pipeline['status']),
            ('Created At', pipeline['created_at'].strftime('%Y-%m-%d %H:%M')),
            ('Started At', pipeline['pipeline_started'].strftime('%Y-%m-%d %H:%M') if pipeline['pipeline_started'] else 'Not started'),
            ('Completed At', pipeline['pipeline_completed'].strftime('%Y-%m-%d %H:%M') if pipeline['pipeline_completed'] else 'Not completed'),
            ('Total Runtime', f"{pipeline['total_runtime_minutes']:.1f} minutes" if pipeline['total_runtime_minutes'] else 'N/A'),
            ('Keywords', pipeline['keyword_count']),
            ('Phases', pipeline['phase_count']),
            ('Regions', len(regions))
        ]
        
        for idx, (label, value) in enumerate(overview_data, start=3):
            ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=idx, column=2, value=value)
        
        # Phase execution details
        phase_start_row = len(overview_data) + 5
        ws.cell(row=phase_start_row, column=1, value='Phase Execution Details').font = Font(bold=True, size=12)
        
        phase_headers = ['Phase', 'Status', 'Started', 'Completed', 'Runtime (min)', 'Details']
        for col, header in enumerate(phase_headers, start=1):
            ws.cell(row=phase_start_row + 1, column=col, value=header).font = Font(bold=True)
        
        for idx, phase in enumerate(phases, start=phase_start_row + 2):
            ws.cell(row=idx, column=1, value=phase['phase_name'])
            ws.cell(row=idx, column=2, value=phase['status'])
            ws.cell(row=idx, column=3, value=phase['started_at'].strftime('%H:%M:%S') if phase['started_at'] else '')
            ws.cell(row=idx, column=4, value=phase['completed_at'].strftime('%H:%M:%S') if phase['completed_at'] else '')
            ws.cell(row=idx, column=5, value=f"{phase['runtime_minutes']:.1f}" if phase['runtime_minutes'] else '')
            
            # Extract key details
            details = ''
            if False:  # Metadata not available in current schema
                meta = {}
                if phase['phase_name'] == 'serp_collection':
                    details = f"Results: {meta.get('serp_results_collected', 0)}"
                elif phase['phase_name'] == 'content_scraping':
                    details = f"Scraped: {meta.get('pages_scraped', 0)}/{meta.get('total_urls', 0)}"
                elif phase['phase_name'] == 'content_analysis':
                    details = f"Analyzed: {meta.get('pages_analyzed', 0)}"
            ws.cell(row=idx, column=6, value=details)
        
        # Keywords section
        keyword_start_row = phase_start_row + len(phases) + 4
        ws.cell(row=keyword_start_row, column=1, value='Keywords').font = Font(bold=True, size=12)
        
        for idx, kw in enumerate(keywords, start=keyword_start_row + 1):
            ws.cell(row=idx, column=1, value=kw['keyword'])
        
        self._auto_adjust_columns(ws)
