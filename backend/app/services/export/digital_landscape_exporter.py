"""
Digital Landscape Excel Exporter
Generates comprehensive Excel reports with DSI rankings and analysis data
"""
import io
import json
from datetime import datetime
from typing import Dict, List, Any, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from loguru import logger

from app.core.database import DatabasePool


class DigitalLandscapeExporter:
    """Export digital landscape data to Excel format"""
    
    def __init__(self, db: DatabasePool):
        self.db = db
        
    async def export_pipeline_data(self, pipeline_id: str) -> io.BytesIO:
        """Generate comprehensive Excel export for a pipeline"""
        async with self.db.acquire() as conn:
            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # 1. Company DSI Ranking
            await self._add_company_dsi_sheet(wb, conn, pipeline_id)
            
            # 2. Page Level DSI Ranking
            await self._add_page_dsi_sheet(wb, conn, pipeline_id)
            
            # 3. Full Page Level Data
            await self._add_full_page_data_sheet(wb, conn, pipeline_id)
            
            # 4. SERPs & Page Analysis
            await self._add_serp_analysis_sheet(wb, conn, pipeline_id)
            
            # 5. Video Results
            await self._add_video_results_sheet(wb, conn, pipeline_id)
            
            # 6. News Results
            await self._add_news_results_sheet(wb, conn, pipeline_id)
            
            # 7. Summary Dashboard
            await self._add_summary_sheet(wb, conn, pipeline_id)
            
            # Save to BytesIO
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output
    
    async def _add_company_dsi_sheet(self, wb: Workbook, conn, pipeline_id: str):
        """Add company DSI ranking sheet"""
        ws = wb.create_sheet("Company DSI Rankings")
        
        # Get company DSI data
        companies = await conn.fetch("""
            WITH company_scores AS (
                SELECT 
                    ds.company_domain,
                    ds.dsi_score,
                    ds.keyword_overlap_score,
                    ds.content_relevance_score,
                    ds.market_presence_score,
                    ds.traffic_share_score,
                    ds.metadata,
                    COUNT(DISTINCT sr.keyword_id) as keywords_found,
                    COUNT(DISTINCT sc.url) as pages_scraped,
                    COUNT(DISTINCT oca.id) as pages_analyzed,
                    AVG(sr.position) as avg_position
                FROM dsi_scores ds
                LEFT JOIN serp_results sr ON sr.domain = ds.company_domain 
                    AND sr.pipeline_execution_id = ds.pipeline_execution_id
                LEFT JOIN scraped_content sc ON sc.domain = ds.company_domain
                    AND sc.pipeline_execution_id = ds.pipeline_execution_id
                LEFT JOIN optimized_content_analysis oca ON oca.url = sc.url
                WHERE ds.pipeline_execution_id = $1
                GROUP BY ds.company_domain, ds.dsi_score, ds.keyword_overlap_score,
                         ds.content_relevance_score, ds.market_presence_score,
                         ds.traffic_share_score, ds.metadata
            )
            SELECT *,
                   RANK() OVER (ORDER BY dsi_score DESC) as dsi_rank
            FROM company_scores
            ORDER BY dsi_score DESC
        """, pipeline_id)
        
        # Headers
        headers = [
            "Rank", "Company Domain", "DSI Score", "Keyword Coverage", 
            "Content Relevance", "Market Presence", "Traffic Share",
            "Keywords Found", "Pages Scraped", "Pages Analyzed", 
            "Avg SERP Position", "Formula Type"
        ]
        
        # Style headers
        self._style_headers(ws, headers)
        
        # Add data
        for row_idx, company in enumerate(companies, start=2):
            metadata = company['metadata'] if isinstance(company['metadata'], dict) else {}
            formula_type = metadata.get('formula', 'unknown')
            
            ws.cell(row=row_idx, column=1, value=company['dsi_rank'])
            ws.cell(row=row_idx, column=2, value=company['company_domain'])
            ws.cell(row=row_idx, column=3, value=round(company['dsi_score'], 4))
            ws.cell(row=row_idx, column=4, value=round(company['keyword_overlap_score'], 4))
            ws.cell(row=row_idx, column=5, value=round(company['content_relevance_score'], 4))
            ws.cell(row=row_idx, column=6, value=round(company['market_presence_score'], 4))
            ws.cell(row=row_idx, column=7, value=round(company['traffic_share_score'] or 0, 4))
            ws.cell(row=row_idx, column=8, value=company['keywords_found'])
            ws.cell(row=row_idx, column=9, value=company['pages_scraped'])
            ws.cell(row=row_idx, column=10, value=company['pages_analyzed'])
            ws.cell(row=row_idx, column=11, value=round(float(company['avg_position']) if company['avg_position'] else 0, 1))
            ws.cell(row=row_idx, column=12, value=formula_type)
            
            # Color code DSI scores
            self._color_score_cell(ws.cell(row=row_idx, column=3), company['dsi_score'])
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    async def _add_page_dsi_sheet(self, wb: Workbook, conn, pipeline_id: str):
        """Add page-level DSI ranking sheet"""
        ws = wb.create_sheet("Page DSI Rankings")
        
        # Get page-level data with DSI calculations
        pages = await conn.fetch("""
            WITH page_scores AS (
                SELECT 
                    sc.url,
                    sc.domain,
                    sc.title,
                    sr.position,
                    sr.serp_type,
                    k.keyword,
                    ds.dsi_score as company_dsi_score,
                    -- Calculate page-level contribution to DSI
                    CASE 
                        WHEN oda_persona.avg_score IS NOT NULL 
                        THEN oda_persona.avg_score / 10.0
                        ELSE 0.5
                    END as page_relevance_score,
                    oda_persona.avg_score as avg_persona_score,
                    oda_jtbd.avg_score as avg_jtbd_score,
                    oca.overall_sentiment,
                    oca.key_topics
                FROM scraped_content sc
                JOIN serp_results sr ON sr.url = sc.url
                JOIN keywords k ON sr.keyword_id = k.id
                LEFT JOIN dsi_scores ds ON ds.company_domain = sc.domain 
                    AND ds.pipeline_execution_id = sc.pipeline_execution_id
                LEFT JOIN optimized_content_analysis oca ON oca.url = sc.url
                LEFT JOIN (
                    SELECT analysis_id, AVG(CAST(score AS FLOAT)) as avg_score
                    FROM optimized_dimension_analysis
                    WHERE dimension_type = 'persona'
                    GROUP BY analysis_id
                ) oda_persona ON oda_persona.analysis_id = oca.id
                LEFT JOIN (
                    SELECT analysis_id, AVG(CAST(score AS FLOAT)) as avg_score
                    FROM optimized_dimension_analysis
                    WHERE dimension_type = 'jtbd_phase'
                    GROUP BY analysis_id
                ) oda_jtbd ON oda_jtbd.analysis_id = oca.id
                WHERE sc.pipeline_execution_id = $1
                    AND sc.status = 'completed'
            )
            SELECT *,
                   -- Page DSI = Company DSI * Page Relevance
                   company_dsi_score * page_relevance_score as page_dsi_score,
                   RANK() OVER (ORDER BY company_dsi_score * page_relevance_score DESC) as page_rank
            FROM page_scores
            ORDER BY page_dsi_score DESC
        """, pipeline_id)
        
        # Headers
        headers = [
            "Rank", "URL", "Title", "Domain", "Keyword", "SERP Type", "Position",
            "Page DSI Score", "Company DSI", "Page Relevance", 
            "Avg Persona Score", "Avg JTBD Score", "Sentiment", "Key Topics"
        ]
        
        self._style_headers(ws, headers)
        
        # Add data
        for row_idx, page in enumerate(pages, start=2):
            ws.cell(row=row_idx, column=1, value=page['page_rank'])
            ws.cell(row=row_idx, column=2, value=page['url'][:255])  # Excel cell limit
            ws.cell(row=row_idx, column=3, value=(page['title'] or '')[:255])
            ws.cell(row=row_idx, column=4, value=page['domain'])
            ws.cell(row=row_idx, column=5, value=page['keyword'])
            ws.cell(row=row_idx, column=6, value=page['serp_type'])
            ws.cell(row=row_idx, column=7, value=page['position'])
            ws.cell(row=row_idx, column=8, value=round(float(page['page_dsi_score'] or 0), 4))
            ws.cell(row=row_idx, column=9, value=round(float(page['company_dsi_score'] or 0), 4))
            ws.cell(row=row_idx, column=10, value=round(float(page['page_relevance_score']), 4))
            ws.cell(row=row_idx, column=11, value=round(float(page['avg_persona_score'] or 0), 2))
            ws.cell(row=row_idx, column=12, value=round(float(page['avg_jtbd_score'] or 0), 2))
            ws.cell(row=row_idx, column=13, value=page['overall_sentiment'] or '')
            ws.cell(row=row_idx, column=14, value=', '.join(page['key_topics'][:3]) if page['key_topics'] else '')
            
            self._color_score_cell(ws.cell(row=row_idx, column=8), float(page['page_dsi_score'] or 0))
        
        self._auto_adjust_columns(ws)
    
    async def _add_full_page_data_sheet(self, wb: Workbook, conn, pipeline_id: str):
        """Add comprehensive page-level data"""
        ws = wb.create_sheet("Full Page Data")
        
        # Get all page data with analysis
        pages = await conn.fetch("""
            SELECT 
                sc.url,
                sc.domain,
                sc.title,
                sc.meta_description,
                LENGTH(sc.content) as content_length,
                sc.status as scrape_status,
                sc.created_at as scraped_at,
                sr.keyword_id,
                k.keyword,
                sr.position,
                sr.serp_type,
                sr.snippet,
                oca.overall_insights,
                oca.overall_sentiment,
                oca.key_topics,
                oca.mentions,
                -- Get all dimension scores as JSON
                (
                    SELECT json_object_agg(
                        dimension_name || '_' || dimension_type, 
                        json_build_object('score', score, 'evidence', key_evidence)
                    )
                    FROM optimized_dimension_analysis
                    WHERE analysis_id = oca.id
                ) as dimension_scores
            FROM scraped_content sc
            JOIN serp_results sr ON sr.url = sc.url AND sr.pipeline_execution_id = sc.pipeline_execution_id
            JOIN keywords k ON sr.keyword_id = k.id
            LEFT JOIN optimized_content_analysis oca ON oca.url = sc.url
            WHERE sc.pipeline_execution_id = $1
            ORDER BY k.keyword, sr.position
        """, pipeline_id)
        
        # Headers
        headers = [
            "URL", "Domain", "Title", "Meta Description", "Content Length",
            "Scrape Status", "Scraped At", "Keyword", "Position", "SERP Type",
            "Snippet", "Overall Insights", "Sentiment", "Key Topics", "Mentions",
            "Lending Leader Score", "Banking Architect Score", "Payments Innovator Score",
            "Problem ID Score", "Solution Explore Score", "Requirements Score",
            "Supplier Select Score", "Validation Score"
        ]
        
        self._style_headers(ws, headers)
        
        # Add data
        for row_idx, page in enumerate(pages, start=2):
            dim_scores = page['dimension_scores'] or {}
            
            ws.cell(row=row_idx, column=1, value=page['url'][:255])
            ws.cell(row=row_idx, column=2, value=page['domain'])
            ws.cell(row=row_idx, column=3, value=(page['title'] or '')[:255])
            ws.cell(row=row_idx, column=4, value=(page['meta_description'] or '')[:255])
            ws.cell(row=row_idx, column=5, value=page['content_length'] or 0)
            ws.cell(row=row_idx, column=6, value=page['scrape_status'])
            ws.cell(row=row_idx, column=7, value=page['scraped_at'].strftime('%Y-%m-%d %H:%M') if page['scraped_at'] else '')
            ws.cell(row=row_idx, column=8, value=page['keyword'])
            ws.cell(row=row_idx, column=9, value=page['position'])
            ws.cell(row=row_idx, column=10, value=page['serp_type'])
            ws.cell(row=row_idx, column=11, value=(page['snippet'] or '')[:255])
            ws.cell(row=row_idx, column=12, value=(page['overall_insights'] or '')[:500])
            ws.cell(row=row_idx, column=13, value=page['overall_sentiment'] or '')
            ws.cell(row=row_idx, column=14, value=', '.join(page['key_topics'][:5]) if page['key_topics'] else '')
            ws.cell(row=row_idx, column=15, value=json.dumps(page['mentions'])[:255] if page['mentions'] else '')
            
            # Persona scores
            ws.cell(row=row_idx, column=16, value=self._get_dimension_score(dim_scores, 'The Modern Lending Leader_persona'))
            ws.cell(row=row_idx, column=17, value=self._get_dimension_score(dim_scores, 'The Digital Banking Architect_persona'))
            ws.cell(row=row_idx, column=18, value=self._get_dimension_score(dim_scores, 'The Payments Innovator_persona'))
            
            # JTBD scores
            ws.cell(row=row_idx, column=19, value=self._get_dimension_score(dim_scores, 'Problem Identification_jtbd_phase'))
            ws.cell(row=row_idx, column=20, value=self._get_dimension_score(dim_scores, 'Solution Exploration_jtbd_phase'))
            ws.cell(row=row_idx, column=21, value=self._get_dimension_score(dim_scores, 'Requirements Building_jtbd_phase'))
            ws.cell(row=row_idx, column=22, value=self._get_dimension_score(dim_scores, 'Supplier Selection_jtbd_phase'))
            ws.cell(row=row_idx, column=23, value=self._get_dimension_score(dim_scores, 'Validation_jtbd_phase'))
        
        self._auto_adjust_columns(ws)
    
    async def _add_serp_analysis_sheet(self, wb: Workbook, conn, pipeline_id: str):
        """Add SERP analysis sheet"""
        ws = wb.create_sheet("SERP Analysis")
        
        # Get SERP analysis data
        serp_data = await conn.fetch("""
            WITH serp_analysis AS (
                SELECT 
                    k.keyword,
                    sr.serp_type,
                    sr.position,
                    sr.url,
                    sr.domain,
                    sr.title,
                    sr.snippet,
                    sc.status as scrape_status,
                    oca.id IS NOT NULL as is_analyzed,
                    ds.dsi_score
                FROM serp_results sr
                JOIN keywords k ON sr.keyword_id = k.id
                LEFT JOIN scraped_content sc ON sc.url = sr.url 
                    AND sc.pipeline_execution_id = sr.pipeline_execution_id
                LEFT JOIN optimized_content_analysis oca ON oca.url = sr.url
                LEFT JOIN dsi_scores ds ON ds.company_domain = sr.domain
                    AND ds.pipeline_execution_id = sr.pipeline_execution_id
                WHERE sr.pipeline_execution_id = $1
            )
            SELECT *,
                   RANK() OVER (PARTITION BY keyword ORDER BY position) as keyword_rank
            FROM serp_analysis
            ORDER BY keyword, position
        """, pipeline_id)
        
        # Headers
        headers = [
            "Keyword", "Rank", "SERP Type", "Position", "URL", "Domain",
            "Title", "Snippet", "Scrape Status", "Analyzed", "Company DSI Score"
        ]
        
        self._style_headers(ws, headers)
        
        # Add data
        for row_idx, serp in enumerate(serp_data, start=2):
            ws.cell(row=row_idx, column=1, value=serp['keyword'])
            ws.cell(row=row_idx, column=2, value=serp['keyword_rank'])
            ws.cell(row=row_idx, column=3, value=serp['serp_type'])
            ws.cell(row=row_idx, column=4, value=serp['position'])
            ws.cell(row=row_idx, column=5, value=serp['url'][:255])
            ws.cell(row=row_idx, column=6, value=serp['domain'])
            ws.cell(row=row_idx, column=7, value=(serp['title'] or '')[:255])
            ws.cell(row=row_idx, column=8, value=(serp['snippet'] or '')[:500])
            ws.cell(row=row_idx, column=9, value=serp['scrape_status'] or 'Not scraped')
            ws.cell(row=row_idx, column=10, value='Yes' if serp['is_analyzed'] else 'No')
            ws.cell(row=row_idx, column=11, value=round(float(serp['dsi_score']), 4) if serp['dsi_score'] else '')
            
            # Highlight top positions
            if serp['position'] <= 3:
                ws.cell(row=row_idx, column=4).fill = PatternFill(
                    start_color="90EE90", end_color="90EE90", fill_type="solid"
                )
        
        self._auto_adjust_columns(ws)
    
    async def _add_video_results_sheet(self, wb: Workbook, conn, pipeline_id: str):
        """Add video results sheet"""
        ws = wb.create_sheet("Video Results")
        
        # Get video results
        videos = await conn.fetch("""
            SELECT 
                sr.url,
                sr.domain,
                sr.title,
                sr.snippet,
                sr.position,
                k.keyword,
                sc.status as scrape_status,
                ds.dsi_score,
                yc.channel_id,
                yc.channel_name,
                yc.company_domain as resolved_company
            FROM serp_results sr
            JOIN keywords k ON sr.keyword_id = k.id
            LEFT JOIN scraped_content sc ON sc.url = sr.url 
                AND sc.pipeline_execution_id = sr.pipeline_execution_id
            LEFT JOIN dsi_scores ds ON ds.company_domain = sr.domain
                AND ds.pipeline_execution_id = sr.pipeline_execution_id
            LEFT JOIN youtube_channels yc ON yc.channel_url = sr.url
            WHERE sr.pipeline_execution_id = $1
                AND sr.serp_type = 'video'
            ORDER BY k.keyword, sr.position
        """, pipeline_id)
        
        # Headers
        headers = [
            "Keyword", "Position", "URL", "Domain", "Title", "Description",
            "DSI Score", "Channel ID", "Channel Name", "Resolved Company",
            "Scrape Status"
        ]
        
        self._style_headers(ws, headers)
        
        # Add data
        for row_idx, video in enumerate(videos, start=2):
            ws.cell(row=row_idx, column=1, value=video['keyword'])
            ws.cell(row=row_idx, column=2, value=video['position'])
            ws.cell(row=row_idx, column=3, value=video['url'][:255])
            ws.cell(row=row_idx, column=4, value=video['domain'])
            ws.cell(row=row_idx, column=5, value=(video['title'] or '')[:255])
            ws.cell(row=row_idx, column=6, value=(video['snippet'] or '')[:500])
            ws.cell(row=row_idx, column=7, value=round(float(video['dsi_score']), 4) if video['dsi_score'] else '')
            ws.cell(row=row_idx, column=8, value=video['channel_id'] or '')
            ws.cell(row=row_idx, column=9, value=video['channel_name'] or '')
            ws.cell(row=row_idx, column=10, value=video['resolved_company'] or '')
            ws.cell(row=row_idx, column=11, value=video['scrape_status'] or 'Not scraped')
        
        self._auto_adjust_columns(ws)
    
    async def _add_news_results_sheet(self, wb: Workbook, conn, pipeline_id: str):
        """Add news results sheet"""
        ws = wb.create_sheet("News Results")
        
        # Get news results
        news = await conn.fetch("""
            SELECT 
                sr.url,
                sr.domain,
                sr.title,
                sr.snippet,
                sr.position,
                k.keyword,
                sr.search_date,
                sc.status as scrape_status,
                sc.content IS NOT NULL as has_content,
                oca.overall_sentiment,
                oca.key_topics,
                ds.dsi_score
            FROM serp_results sr
            JOIN keywords k ON sr.keyword_id = k.id
            LEFT JOIN scraped_content sc ON sc.url = sr.url 
                AND sc.pipeline_execution_id = sr.pipeline_execution_id
            LEFT JOIN optimized_content_analysis oca ON oca.url = sr.url
            LEFT JOIN dsi_scores ds ON ds.company_domain = sr.domain
                AND ds.pipeline_execution_id = sr.pipeline_execution_id
            WHERE sr.pipeline_execution_id = $1
                AND sr.serp_type = 'news'
            ORDER BY sr.search_date DESC, k.keyword, sr.position
        """, pipeline_id)
        
        # Headers
        headers = [
            "Date", "Keyword", "Position", "URL", "Domain", "Title", "Snippet",
            "DSI Score", "Scrape Status", "Has Content", "Sentiment", "Key Topics"
        ]
        
        self._style_headers(ws, headers)
        
        # Add data
        for row_idx, article in enumerate(news, start=2):
            ws.cell(row=row_idx, column=1, value=article['search_date'].strftime('%Y-%m-%d'))
            ws.cell(row=row_idx, column=2, value=article['keyword'])
            ws.cell(row=row_idx, column=3, value=article['position'])
            ws.cell(row=row_idx, column=4, value=article['url'][:255])
            ws.cell(row=row_idx, column=5, value=article['domain'])
            ws.cell(row=row_idx, column=6, value=(article['title'] or '')[:255])
            ws.cell(row=row_idx, column=7, value=(article['snippet'] or '')[:500])
            ws.cell(row=row_idx, column=8, value=round(float(article['dsi_score']), 4) if article['dsi_score'] else '')
            ws.cell(row=row_idx, column=9, value=article['scrape_status'] or 'Not scraped')
            ws.cell(row=row_idx, column=10, value='Yes' if article['has_content'] else 'No')
            ws.cell(row=row_idx, column=11, value=article['overall_sentiment'] or '')
            ws.cell(row=row_idx, column=12, value=', '.join(article['key_topics'][:3]) if article['key_topics'] else '')
        
        self._auto_adjust_columns(ws)
    
    async def _add_summary_sheet(self, wb: Workbook, conn, pipeline_id: str):
        """Add summary dashboard as first sheet"""
        ws = wb.create_sheet("Summary Dashboard", 0)
        
        # Get summary statistics
        stats = await conn.fetchrow("""
            WITH pipeline_stats AS (
                SELECT 
                    pe.created_at,
                    pe.completed_at,
                    pe.status,
                    (SELECT COUNT(DISTINCT keyword_id) FROM serp_results WHERE pipeline_execution_id = $1) as total_keywords,
                    (SELECT COUNT(*) FROM serp_results WHERE pipeline_execution_id = $1) as total_serp_results,
                    (SELECT COUNT(*) FROM serp_results WHERE pipeline_execution_id = $1 AND serp_type = 'organic') as organic_results,
                    (SELECT COUNT(*) FROM serp_results WHERE pipeline_execution_id = $1 AND serp_type = 'news') as news_results,
                    (SELECT COUNT(*) FROM serp_results WHERE pipeline_execution_id = $1 AND serp_type = 'video') as video_results,
                    (SELECT COUNT(DISTINCT domain) FROM serp_results WHERE pipeline_execution_id = $1) as unique_domains,
                    (SELECT COUNT(*) FROM scraped_content WHERE pipeline_execution_id = $1) as total_scraped,
                    (SELECT COUNT(*) FROM scraped_content WHERE pipeline_execution_id = $1 AND status = 'completed') as successful_scrapes,
                    (SELECT COUNT(DISTINCT oca.url) FROM optimized_content_analysis oca 
                     JOIN scraped_content sc ON oca.url = sc.url 
                     WHERE sc.pipeline_execution_id = $1) as total_analyzed,
                    (SELECT COUNT(*) FROM dsi_scores WHERE pipeline_execution_id = $1) as companies_scored,
                    (SELECT AVG(dsi_score) FROM dsi_scores WHERE pipeline_execution_id = $1) as avg_dsi_score,
                    (SELECT MAX(dsi_score) FROM dsi_scores WHERE pipeline_execution_id = $1) as max_dsi_score
                FROM pipeline_executions pe
                WHERE pe.id = $1
            )
            SELECT * FROM pipeline_stats
        """, pipeline_id)
        
        # Title
        ws['A1'] = "Digital Landscape Analysis Report"
        ws['A1'].font = Font(size=18, bold=True)
        ws['A2'] = f"Pipeline ID: {pipeline_id}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Pipeline Info
        ws['A5'] = "Pipeline Information"
        ws['A5'].font = Font(size=14, bold=True)
        
        info_data = [
            ['Status:', stats['status']],
            ['Started:', stats['created_at'].strftime('%Y-%m-%d %H:%M') if stats['created_at'] else 'N/A'],
            ['Completed:', stats['completed_at'].strftime('%Y-%m-%d %H:%M') if stats['completed_at'] else 'In Progress'],
            ['Duration:', self._calculate_duration(stats['created_at'], stats['completed_at'])],
        ]
        
        for idx, (label, value) in enumerate(info_data, start=6):
            ws[f'A{idx}'] = label
            ws[f'B{idx}'] = value
            ws[f'A{idx}'].font = Font(bold=True)
        
        # Key Metrics
        ws['D5'] = "Key Metrics"
        ws['D5'].font = Font(size=14, bold=True)
        
        metrics_data = [
            ['Total Keywords:', stats['total_keywords']],
            ['SERP Results:', stats['total_serp_results']],
            ['Unique Domains:', stats['unique_domains']],
            ['Pages Scraped:', f"{stats['successful_scrapes']}/{stats['total_scraped']} ({stats['successful_scrapes']/stats['total_scraped']*100:.1f}%)"],
            ['Pages Analyzed:', stats['total_analyzed']],
            ['Companies Scored:', stats['companies_scored']],
            ['Average DSI Score:', f"{float(stats['avg_dsi_score'] or 0):.3f}"],
            ['Top DSI Score:', f"{float(stats['max_dsi_score'] or 0):.3f}"],
        ]
        
        for idx, (label, value) in enumerate(metrics_data, start=6):
            ws[f'D{idx}'] = label
            ws[f'E{idx}'] = value
            ws[f'D{idx}'].font = Font(bold=True)
        
        # SERP Breakdown
        ws['A15'] = "SERP Results Breakdown"
        ws['A15'].font = Font(size=14, bold=True)
        
        serp_headers = ['Type', 'Count', 'Percentage']
        for col_idx, header in enumerate(serp_headers, start=1):
            cell = ws.cell(row=16, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        serp_data = [
            ['Organic', stats['organic_results'], f"{stats['organic_results']/stats['total_serp_results']*100:.1f}%"],
            ['News', stats['news_results'], f"{stats['news_results']/stats['total_serp_results']*100:.1f}%"],
            ['Video', stats['video_results'], f"{stats['video_results']/stats['total_serp_results']*100:.1f}%"],
        ]
        
        for row_idx, row_data in enumerate(serp_data, start=17):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Top Companies by DSI
        ws['D15'] = "Top 10 Companies by DSI Score"
        ws['D15'].font = Font(size=14, bold=True)
        
        top_companies = await conn.fetch("""
            SELECT company_domain, dsi_score
            FROM dsi_scores
            WHERE pipeline_execution_id = $1
            ORDER BY dsi_score DESC
            LIMIT 10
        """, pipeline_id)
        
        company_headers = ['Rank', 'Company', 'DSI Score']
        for col_idx, header in enumerate(company_headers, start=4):
            cell = ws.cell(row=16, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for row_idx, company in enumerate(top_companies, start=17):
            ws.cell(row=row_idx, column=4, value=row_idx - 16)
            ws.cell(row=row_idx, column=5, value=company['company_domain'])
            ws.cell(row=row_idx, column=6, value=round(float(company['dsi_score']), 4))
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def _style_headers(self, ws, headers: List[str]):
        """Apply consistent header styling"""
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    def _color_score_cell(self, cell, score: float):
        """Color code score cells"""
        if score >= 0.7:
            color = "90EE90"  # Light green
        elif score >= 0.4:
            color = "FFFFE0"  # Light yellow
        else:
            color = "FFB6C1"  # Light red
        
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _get_dimension_score(self, dim_scores: Dict, dimension_key: str) -> float:
        """Extract dimension score from JSON"""
        if not dim_scores:
            return 0
        
        # Handle case where dim_scores might be JSON string
        if isinstance(dim_scores, str):
            try:
                import json
                dim_scores = json.loads(dim_scores)
            except:
                return 0
        
        if dimension_key in dim_scores:
            score_data = dim_scores[dimension_key]
            if isinstance(score_data, dict):
                return float(score_data.get('score', 0))
            else:
                return float(score_data or 0)
        return 0
    
    def _calculate_duration(self, start_time, end_time) -> str:
        """Calculate duration between timestamps"""
        if not start_time:
            return "N/A"
        
        end = end_time or datetime.now(start_time.tzinfo)
        duration = end - start_time
        
        hours = duration.total_seconds() / 3600
        if hours < 1:
            return f"{int(duration.total_seconds() / 60)} minutes"
        elif hours < 24:
            return f"{hours:.1f} hours"
        else:
            return f"{hours / 24:.1f} days"
